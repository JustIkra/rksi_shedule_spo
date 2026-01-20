"""
Excel export service for Events Portal.

Provides functionality to export events and categories to Excel files.
"""

from io import BytesIO

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..models import Category, Event


async def export_excel(db: AsyncSession) -> bytes:
    """
    Export all events and categories to an Excel file.

    Generates an Excel workbook with all categories and their events.
    Format:
    - Each month/category as a section
    - Columns: Number, Name, Date, Responsible, Location, Description
    - Formatted with headers and basic styling

    Args:
        db: Async database session.

    Returns:
        Bytes of the generated Excel file (.xlsx format).

    Raises:
        Exception: If export fails.

    TODO: Implement actual Excel generation logic using openpyxl.
    """
    # Fetch all categories with events and photos
    result = await db.execute(
        select(Category)
        .options(
            selectinload(Category.events).selectinload(Event.photos)
        )
        .order_by(Category.sort_order, Category.month)
    )
    categories = result.scalars().all()

    # TODO: Generate actual Excel file
    # Example implementation outline:
    # 1. Create workbook with openpyxl
    # 2. Create worksheet
    # 3. Add headers with styling
    # 4. For each category:
    #    - Add category name as section header
    #    - Add all events in the category
    # 5. Auto-size columns
    # 6. Save to BytesIO and return bytes

    # Stub implementation - returns minimal valid xlsx
    # This creates a minimal Excel file that can be opened
    output = BytesIO()

    try:
        # Try to use openpyxl if available
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Events"

        # Headers
        headers = ["Number", "Name", "Date", "Responsible", "Location", "Description", "Category", "Photos"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data
        row = 2
        for category in categories:
            for event in category.events:
                ws.cell(row=row, column=1, value=event.number)
                ws.cell(row=row, column=2, value=event.name)
                ws.cell(row=row, column=3, value=event.event_date)
                ws.cell(row=row, column=4, value=event.responsible)
                ws.cell(row=row, column=5, value=event.location)
                ws.cell(row=row, column=6, value=event.description)
                ws.cell(row=row, column=7, value=category.name)
                # Photos column - URLs separated by comma
                photo_urls = ", ".join(
                    f"/uploads/{photo.original_path}" for photo in event.photos
                ) if event.photos else ""
                ws.cell(row=row, column=8, value=photo_urls)
                row += 1

        # Auto-size columns (approximate)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output)

    except ImportError:
        # Fallback: return empty bytes with warning
        # In production, openpyxl should be installed
        raise Exception(
            "openpyxl is not installed. Please install it: pip install openpyxl"
        )

    output.seek(0)
    return output.read()
