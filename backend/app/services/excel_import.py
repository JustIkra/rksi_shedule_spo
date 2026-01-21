"""
Excel import service for Events Portal.

Provides functionality to import events and categories from Excel files.
"""

import os
import re
import shutil
from io import BytesIO
from pathlib import Path

from fastapi import UploadFile
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..config import get_settings
from ..models import Category, Event, Photo


def build_merged_values_map(ws):
    """
    Build a mapping of (row, col) -> value for all merged cell ranges.

    In Excel merged cells, only the top-left cell contains the value.
    This function distributes that value to all cells in the merged range.
    """
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_left_value
    return merged_values


def get_cell_value(ws, row, col, merged_values):
    """Get cell value, considering merged cell ranges."""
    if (row, col) in merged_values:
        return merged_values[(row, col)]
    return ws.cell(row, col).value


def _get_cell_background_color(ws, row, col) -> str | None:
    """Return the RGB hex color (uppercase) for a cell's background, if set."""
    cell = ws.cell(row, col)
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None

    rgb = fill.fgColor.rgb
    if rgb and rgb != "00000000":
        return rgb.upper()
    return None


def _row_has_merged_cells(ws, row_idx: int) -> bool:
    """Check if the given row participates in any merged range."""
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row_idx <= merged_range.max_row:
            return True
    return False


def _is_empty_value(val) -> bool:
    """Consider None or blank/whitespace-only strings as empty."""
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == '':
        return True
    return False


def _is_event_number(value) -> bool:
    """
    Check if value is an event number (numeric like 1, 2, 24, 119).

    Returns False for category headers like "1. Заседания Президиума"
    or text-only values.
    """
    if value is None:
        return False

    str_val = str(value).strip()

    # Pure number (int or float that's whole)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return False
        return True

    # String that is just a number
    if re.match(r'^\d+$', str_val):
        return True

    return False


def _is_category_row(ws, row_idx, merged_values) -> bool:
    """
    Determine if a row is a category header.

    Category row criteria:
    - Cell A has yellow background (00FFF2CC)
    - Column A contains non-numeric text
    - Columns C, D, E are empty
    - Row may be merged (optional, supportive)
    """
    color = _get_cell_background_color(ws, row_idx, 1)
    if color != "00FFF2CC":
        return False

    col_a = get_cell_value(ws, row_idx, 1, merged_values)
    has_merge = _row_has_merged_cells(ws, row_idx)
    if has_merge:
        # When a row is merged across columns, merged_values will duplicate the text;
        # treat C/D/E as empty for category detection.
        col_c = col_d = col_e = None
    else:
        col_c = get_cell_value(ws, row_idx, 3, merged_values)
        col_d = get_cell_value(ws, row_idx, 4, merged_values)
        col_e = get_cell_value(ws, row_idx, 5, merged_values)

    # Must have a non-numeric value in column A
    if _is_empty_value(col_a) or _is_event_number(col_a):
        return False

    # Columns C/D/E must be empty
    if not (_is_empty_value(col_c) and _is_empty_value(col_d) and _is_empty_value(col_e)):
        return False

    # Optional: merged rows are a common category marker; not required
    return True


def _is_event_row(first_col, second_col) -> bool:
    """
    Determine if a row is an event.

    Event row criteria:
    - Column A is a number (1, 2, 24...)
    - OR Column B has text (event without number)
    """
    # If A is a number - it's an event
    if _is_event_number(first_col):
        return True

    # If B has text but A is not a category marker - it's an event without number
    if second_col and not first_col:
        return True

    return False


async def import_excel(file: UploadFile, db: AsyncSession) -> dict:
    """
    Import events and categories from an Excel file.

    Parses the Excel file with 12 sheets (one per month) and creates
    categories and events in the database.

    Expected Excel format:
    - 12 sheets named by months (Январь-Декабрь)
    - Row 1: Header row (skipped)
    - Category rows: Column A has text (not a number) or is empty with text in B
    - Event rows: Column A has a number (1, 2, 24...)

    Columns:
    - A: Event number OR category name
    - B: Event name (or empty for category-in-A format)
    - C: Event date
    - D: Responsible person/department
    - E: Location
    - F: Description (Пояснение)
    - G: Links (ignored during import)
    - H: Photos (ignored during import)

    Args:
        file: Uploaded Excel file (UploadFile from FastAPI).
        db: Async database session.

    Returns:
        Dictionary with import statistics.

    Raises:
        ValueError: If file format is invalid or cannot be parsed.
    """
    from openpyxl import load_workbook

    # Read file content
    content = await file.read()
    await file.seek(0)

    if len(content) == 0:
        raise ValueError("Empty file provided")

    try:
        wb = load_workbook(BytesIO(content))
    except Exception as e:
        raise ValueError(f"Cannot parse Excel file: {str(e)}")

    months_ru = [
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
        'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    ]

    imported_events = 0
    imported_categories = 0
    errors = []

    # Get settings for upload directory
    settings = get_settings()

    # Delete photo files from disk before clearing database
    # Query all photos to get their file paths
    photos_result = await db.execute(select(Photo))
    all_photos = photos_result.scalars().all()

    for photo in all_photos:
        for path_attr in ["original_path", "thumbnail_path"]:
            file_path = getattr(photo, path_attr, None)
            if file_path:
                try:
                    full_path = Path(settings.upload_dir) / file_path
                    if full_path.exists():
                        os.remove(full_path)
                except OSError:
                    # Log error but don't fail the import
                    pass

    # Delete old events and categories (cascade deletes photos too)
    await db.execute(delete(Event))
    await db.execute(delete(Category))

    for month_num, month_name in enumerate(months_ru, 1):
        if month_name not in wb.sheetnames:
            continue

        ws = wb[month_name]
        merged_values = build_merged_values_map(ws)
        current_category = None
        category_sort = 0
        event_sort = 0

        for row_idx in range(2, ws.max_row + 1):
            # Get cell values considering merged cells
            col_a = get_cell_value(ws, row_idx, 1, merged_values)
            col_b = get_cell_value(ws, row_idx, 2, merged_values)
            col_c = get_cell_value(ws, row_idx, 3, merged_values)
            col_d = get_cell_value(ws, row_idx, 4, merged_values)
            col_e = get_cell_value(ws, row_idx, 5, merged_values)
            col_f = get_cell_value(ws, row_idx, 6, merged_values)
            col_g = get_cell_value(ws, row_idx, 7, merged_values)
            col_h = get_cell_value(ws, row_idx, 8, merged_values)

            # Skip completely empty rows
            if _is_empty_value(col_a) and _is_empty_value(col_b):
                continue

            if _is_category_row(ws, row_idx, merged_values):
                category_name = str(col_a).strip() if col_a else ''

                category = Category(
                    name=category_name,
                    month=month_num,
                    sort_order=category_sort
                )
                db.add(category)
                await db.flush()
                current_category = category
                imported_categories += 1
                category_sort += 1
                event_sort = 0  # Reset event sort for new category

            elif _is_event_row(col_a, col_b):
                # This is an event row (has number OR has text in B)
                if not current_category:
                    # Create default category if none exists
                    current_category = Category(
                        name="Общие мероприятия",
                        month=month_num,
                        sort_order=category_sort
                    )
                    db.add(current_category)
                    await db.flush()
                    imported_categories += 1
                    category_sort += 1

                # Handle event number (may be empty for events without number)
                event_number = ''
                if col_a:
                    if isinstance(col_a, float):
                        event_number = str(int(col_a))
                    else:
                        event_number = str(col_a).strip()

                event = Event(
                    category_id=current_category.id,
                    number=event_number,
                    name=str(col_b).strip() if col_b else '',
                    event_date=str(col_c).strip() if col_c else None,
                    responsible=str(col_d).strip() if col_d else None,
                    location=str(col_e).strip() if col_e else None,
                    description=str(col_f).strip() if col_f else None,
                    sort_order=event_sort
                )
                db.add(event)
                imported_events += 1
                event_sort += 1

    await db.commit()

    return {
        "imported_events": imported_events,
        "imported_categories": imported_categories,
        "errors": errors,
    }


async def preview_excel(file: UploadFile) -> dict:
    """
    Preview an Excel file before importing.

    Parses the Excel file and returns statistics without modifying the database.
    Shows counts of categories and events that would be imported,
    plus any warnings about potential issues.

    Args:
        file: Uploaded Excel file (UploadFile from FastAPI).

    Returns:
        Dictionary with preview statistics and warnings.

    Raises:
        ValueError: If file format is invalid or cannot be parsed.
    """
    from openpyxl import load_workbook

    # Read file content
    content = await file.read()
    await file.seek(0)

    if len(content) == 0:
        raise ValueError("Empty file provided")

    try:
        wb = load_workbook(BytesIO(content))
    except Exception as e:
        raise ValueError(f"Cannot parse Excel file: {str(e)}")

    months_ru = [
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
        'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    ]

    total_events = 0
    total_categories = 0
    warnings = []
    months_found = []

    for month_num, month_name in enumerate(months_ru, 1):
        if month_name not in wb.sheetnames:
            continue

        months_found.append(month_name)
        ws = wb[month_name]
        merged_values = build_merged_values_map(ws)
        current_category_name = None

        for row_idx in range(2, ws.max_row + 1):
            col_a = get_cell_value(ws, row_idx, 1, merged_values)
            col_b = get_cell_value(ws, row_idx, 2, merged_values)

            # Skip completely empty rows
            if _is_empty_value(col_a) and _is_empty_value(col_b):
                continue

            if _is_category_row(ws, row_idx, merged_values):
                current_category_name = str(col_a).strip()
                total_categories += 1

            elif _is_event_row(col_a, col_b):
                event_name = str(col_b).strip() if col_b else ''
                if not event_name:
                    event_num = col_a if col_a else '(без номера)'
                    warnings.append(
                        f"{month_name}: Мероприятие #{event_num} без названия"
                    )
                total_events += 1

    if not months_found:
        warnings.append("Не найдено листов с названиями месяцев (Январь-Декабрь)")

    return {
        "categories_count": total_categories,
        "events_count": total_events,
        "months_found": months_found,
        "warnings": warnings,
    }
