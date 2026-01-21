import unittest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.services import excel_import


YELLOW = "00FFF2CC"


def make_ws_with_row(fill_color=None, cols=None, merged=False):
    """Helper to build a worksheet with a single data row at index 2."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["A", "B", "C", "D", "E"])  # header row placeholder
    values = cols or ["", "", "", "", ""]
    ws.append(values)

    if fill_color:
        for cell in ws[2]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    if merged:
        ws.merge_cells("A2:E2")

    merged_values = excel_import.build_merged_values_map(ws)
    return ws, merged_values


class CategoryDetectionTests(unittest.TestCase):
    def test_category_row_detected_with_color_and_empty_cde(self):
        ws, merged_values = make_ws_with_row(fill_color=YELLOW, cols=["Категория", "", None, None, None])
        self.assertTrue(excel_import._is_category_row(ws, 2, merged_values))

    def test_not_category_when_color_missing(self):
        ws, merged_values = make_ws_with_row(fill_color=None, cols=["Категория", "", None, None, None])
        self.assertFalse(excel_import._is_category_row(ws, 2, merged_values))

    def test_not_category_when_data_in_cde(self):
        ws, merged_values = make_ws_with_row(fill_color=YELLOW, cols=["Категория", "", "дата", None, None])
        self.assertFalse(excel_import._is_category_row(ws, 2, merged_values))

    def test_category_detected_with_merged_row(self):
        ws, merged_values = make_ws_with_row(fill_color=YELLOW, cols=["Категория", "", None, None, None], merged=True)
        self.assertTrue(excel_import._is_category_row(ws, 2, merged_values))


if __name__ == "__main__":
    unittest.main()
